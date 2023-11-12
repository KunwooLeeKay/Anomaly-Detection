
import pandas as pd
learning_data = pd.read_excel(r"C:\Users\user\Desktop\LearnPython\Lab\MachineLearningData_.xlsx")

# print(learning_data.head())

feature = ['제조국가', '사용년수', 'C', 'D', 'E', '진동 횟수', '최대 기계압력', '최소 기계압력', '기계 길이',\
                     '기계 무게', '특수 부품 유무', '특수 부품 저항', '일반 부품 저항', '기계 마력', 'O', 'P',\
                     '기계 부품 임피던스', 'P/J', 'S', 'T', 'U', '기계 평균 압력', 'R/J', 'M*J', 'L*J',\
                     'L/J']

target = ['고장 단계', '고장 유무']



X = learning_data[feature]

Y1 = learning_data[target[0]] # 1번이 고장단계, 2번이 고장 유무
Y2 = learning_data[target[1]]

import random
from openpyxl import Workbook

from sklearn.model_selection import train_test_split

from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression

from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression

from sklearn.metrics import mean_squared_error
from sklearn.metrics import confusion_matrix

def sampler(num):
  feature_sample_list = []
  while(True):
    feature = ['제조국가', '사용년수', 'C', 'D', 'E', '진동 횟수', '최대 기계압력', '최소 기계압력', '기계 길이',\
                     '기계 무게', '특수 부품 유무', '특수 부품 저항', '일반 부품 저항', '기계 마력', 'O', 'P',\
                     '기계 부품 임피던스', 'P/J', 'S', 'T', 'U', '기계 평균 압력', 'R/J', 'M*J', 'L*J',\
                     'L/J']
    feature = random.sample(feature, num)
    feature = sorted(feature)
    flag = True
    for i in range(0, len(feature_sample_list)):
      if feature == feature_sample_list[i]:
        flag = False
    if flag == True:
      feature_sample_list.append(feature)
    if len(feature_sample_list) == 1000:
      break
  return feature_sample_list

for num in [5, 10, 15, 17]:
  wb = Workbook()
  ws = wb.active
  ws.cell(row = 1, column = 1,value = "used feature")
  ws.cell(row = 1, column = 2,value = "RandomForestRegressor 고장단계 예측 정확도(MSE)")
  ws.cell(row = 1, column = 3,value = "LinearRegression 고장단계 예측 정확도(MSE)")
  ws.cell(row = 1, column = 4,value = "LogisticRegression (1-sensitivity)")
  ws.cell(row = 1, column = 5,value = "LogisticRegression (1-specificity)")
  ws.cell(row = 1, column = 6,value = "RandomForestClassifier (1-sensitivity)")
  ws.cell(row = 1, column = 7,value = "RandomForestClassifier (1-specificity)")

  feature_list = sampler(num)

  for j in range(0,1000):
    feature = feature_list[j]

    str_feature = ''
    for i in range(0, len(feature)):
      str_feature += feature[i]
      if i != len(feature) - 1:
        str_feature += ' , '

    ws.cell(row = j+2, column = 1, value = str_feature)

    try:
      X = learning_data[feature]

      Y1 = learning_data[target[0]] # 1번이 고장단계, 2번이 고장 유무
      Y2 = learning_data[target[1]]

      # 고장단계 예측

      # Random Forest Regressor 

      train_X, val_X, train_Y, val_Y = train_test_split(X, Y1, random_state = 1, test_size = 0.2) # 고장 단계 예측
      learn_model = RandomForestRegressor(random_state = 1)
      learn_model.fit(train_X, train_Y)

      val_predictions = learn_model.predict(val_X)
      ws.cell(row = j+2, column = 2, value = mean_squared_error(val_Y, val_predictions))
              
      # # Linear Regression Model
      
      # learn_model = LinearRegression()
      # learn_model.fit(train_X, train_Y)
      # val_predictions = learn_model.predict(val_X)
      # ws.cell(row = j+2, column = 3, value = mean_squared_error(val_Y, val_predictions))

      # 고장유무 예측

      train_X, val_X, train_Y, val_Y = train_test_split(X, Y2, random_state = 1, test_size = 0.2)  # 고장 유무 예측

      # Logistic Regression Model
      
      # learn_model = LogisticRegression(random_state = 1)
      # learn_model.fit(train_X, train_Y)

      # val_predictions = learn_model.predict(val_X)

      # tn, fp, fn, tp = confusion_matrix(val_Y, val_predictions).ravel()
      # specificity = tn / (tn + fp)
      # sensitivity = tp / (tp + fn)
      # ws.cell(row = j+2, column = 4, value = 1 - sensitivity)
      # ws.cell(row = j+2, column = 5, value = 1 - specificity)
    
      # Random Forest Classifier

      learn_model = RandomForestClassifier(random_state = 1)
      learn_model.fit(train_X, train_Y)

      val_predictions = learn_model.predict(val_X)

      tn, fp, fn, tp = confusion_matrix(val_Y, val_predictions).ravel()
      specificity = tn / (tn + fp)
      sensitivity = tp / (tp + fn)

      ws.cell(row = j+2, column = 6, value = 1 - sensitivity)
      ws.cell(row = j+2, column = 7, value = 1 - specificity)
    except:
      ws.cell(row = j+2, column = 8, value = "Error")

    wb.save("/content/drive/MyDrive/Results"+str(num)+"features.xlsx")
    wb.close()