from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\user\Desktop\고장진단 프로젝트\2.24발표\find_optimal.xlsx")
ws = wb.active

accuracy1 = 0
recall1 = 0
precision1 = 0
accuracy2 = 0
recall2 = 0
precision2 = 0
count1 = 0
count2 = 0


for col in range(2, ws.max_column):
    if col%2 == 0:
        accuracy1 += ws.cell(row = 5, column=col).value
        recall1 += ws.cell(row = 6, column=col).value
        precision1 += ws.cell(row = 7, column=col).value
        count1 += 1
    else:
        accuracy2 += ws.cell(row = 5, column=col).value 
        recall2 += ws.cell(row = 6, column=col).value
        precision2 += ws.cell(row = 7, column=col).value
        count2 += 1

print("One-Hot : accuracy {}, recall {}, precision {}".format(accuracy1/count1, recall1/count1, precision1/count1))
print("NO One-Hot : accuracy {}, recall {}, precision {}".format(accuracy2/count2, recall2/count2, precision2/count2))

wb.close()